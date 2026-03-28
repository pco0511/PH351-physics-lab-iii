#!/usr/bin/env python3
"""
Convert Hall effect experiment raw data files to HDF5 format.

Raw data layout
---------------
data/raw/
    current-Bfield/     Excel file: coil current -> magnetic field calibration
    longitudinal/       Longitudinal voltage time-series per coil current
        Ge/             Intrinsic (undoped) Ge
        Ge-n/           n-type Ge
        Ge-p/           p-type Ge
    transverse/         Transverse (Hall) voltage time-series per coil current
        Ge/
        Ge-n/
        Ge-p/

Output HDF5 files
-----------------
data/hdf5/
    calibration.h5          Coil current <-> B-field calibration
    voltage_measurements.h5 Sample voltage vs coil current for each sample

HDF5 group hierarchy mirrors the raw directory tree.
Measurement metadata from file headers is stored as group/dataset attributes.
"""

import re
import sys
from pathlib import Path

import h5py
import numpy as np

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
ROOT = Path(__file__).resolve().parent.parent
RAW_DIR = ROOT / "data" / "raw"
HDF5_DIR = ROOT / "data" / "hdf5"

CALIB_H5 = HDF5_DIR / "calibration.h5"
VOLTAGE_H5 = HDF5_DIR / "voltage_measurements.h5"

SAMPLE_TYPES = {
    "Ge":   {"description": "Intrinsic (undoped) germanium", "doping": "intrinsic"},
    "Ge-n": {"description": "n-type doped germanium",        "doping": "n-type"},
    "Ge-p": {"description": "p-type doped germanium",        "doping": "p-type"},
}


# ---------------------------------------------------------------------------
# txt file parsing
# ---------------------------------------------------------------------------

def _parse_def_column(token: str) -> dict:
    """Parse a single DEF column token like: "Time" t / s"""
    m = re.match(r'"([^"]+)"\s+(\S+)\s*/\s*(.+)', token.strip())
    if m:
        return {"label": m.group(1), "varname": m.group(2), "unit": m.group(3).strip()}
    return {"label": token.strip(), "varname": token.strip(), "unit": ""}


def parse_txt_file(filepath: Path) -> tuple[dict, np.ndarray]:
    """
    Parse a Sensor & Software / Phywe-style tab-separated measurement file.

    Header format (lines 1-5):
        MIN=<v0>  <v1>  <v2>
        MAX=<v0>  <v1>  <v2>
        SCALE=<v0>  <v1>  <v2>
        DEC=<v0>  <v1>  <v2>
        DEF="Label0" var0 / unit0  "Label1" var1 / unit1  ...

    Data (line 6 onward):
        <t>  <col1>  <col2>  ...  (tab-separated, trailing columns may be absent)

    Returns
    -------
    meta : dict
        keys: MIN, MAX, SCALE, DEC (lists of str), columns (list of dicts)
    data : np.ndarray, shape (N, n_cols)
        NaN where values are missing.
    """
    with open(filepath, "r", encoding="utf-8", errors="replace") as fh:
        lines = fh.readlines()

    if len(lines) < 6:
        raise ValueError(f"File too short: {filepath}")

    # ---- header ----
    meta: dict = {}
    key_rows = ["MIN", "MAX", "SCALE", "DEC"]
    for i, key in enumerate(key_rows):
        parts = lines[i].rstrip("\n").split("\t")
        # first token is  KEY=value
        first_val = parts[0].split("=", 1)[1]
        meta[key] = [first_val] + parts[1:]

    # DEF line
    def_line = lines[4].rstrip("\n")
    # Strip leading "DEF=" then split on tab
    def_line = re.sub(r"^DEF=", "", def_line)
    def_tokens = def_line.split("\t")
    meta["columns"] = [_parse_def_column(tok) for tok in def_tokens]
    n_cols = len(meta["columns"])

    # ---- data ----
    rows = []
    for line in lines[5:]:
        line = line.rstrip("\n")
        if not line.strip():
            continue
        parts = line.split("\t")
        try:
            row = []
            for j in range(n_cols):
                val = parts[j].strip() if j < len(parts) else ""
                row.append(float(val) if val else np.nan)
            rows.append(row)
        except ValueError:
            continue

    data = np.array(rows, dtype=float) if rows else np.empty((0, n_cols))
    return meta, data


def coil_current_A(filename: str) -> float:
    """Convert filename stem (e.g. '32') to coil current in Amperes (3.2 A)."""
    return int(Path(filename).stem) / 10.0


# ---------------------------------------------------------------------------
# HDF5 helpers
# ---------------------------------------------------------------------------

def set_attrs(obj: h5py.HLObject, **kwargs):
    """Write keyword arguments as HDF5 attributes, skipping None values."""
    for k, v in kwargs.items():
        if v is not None:
            obj.attrs[k] = v


def write_dataset(group: h5py.Group, name: str, data: np.ndarray,
                  unit: str = "", description: str = ""):
    ds = group.create_dataset(name, data=data, compression="gzip",
                              compression_opts=4)
    set_attrs(ds, unit=unit, description=description)
    return ds


# ---------------------------------------------------------------------------
# Voltage-measurement HDF5 builder
# ---------------------------------------------------------------------------

def build_voltage_h5():
    """Convert longitudinal/ and transverse/ txt files into voltage_measurements.h5."""

    HDF5_DIR.mkdir(parents=True, exist_ok=True)

    meas_dirs = {
        "longitudinal": RAW_DIR / "longitudinal",
        "transverse":   RAW_DIR / "transverse",
    }

    with h5py.File(VOLTAGE_H5, "w") as hf:
        set_attrs(hf,
                  description="Hall effect experiment: voltage measurements on Ge samples",
                  experiment="Hall effect",
                  sample_material="Germanium",
                  source_directory=str(RAW_DIR))

        for meas_type, meas_dir in meas_dirs.items():
            if not meas_dir.exists():
                print(f"  [SKIP] {meas_dir} not found", file=sys.stderr)
                continue

            meas_grp = hf.require_group(meas_type)
            set_attrs(meas_grp,
                      description=f"{meas_type.capitalize()} voltage measurements",
                      voltage_type=meas_type)

            for sample_dir in sorted(meas_dir.iterdir()):
                if not sample_dir.is_dir():
                    continue
                sample_name = sample_dir.name
                if sample_name not in SAMPLE_TYPES:
                    print(f"  [WARN] Unknown sample dir '{sample_name}', skipping",
                          file=sys.stderr)
                    continue

                sample_grp = meas_grp.require_group(sample_name)
                set_attrs(sample_grp, **SAMPLE_TYPES[sample_name])

                txt_files = sorted(sample_dir.glob("*.txt"),
                                   key=lambda p: int(p.stem))

                for txt_path in txt_files:
                    I_coil = coil_current_A(txt_path.name)
                    grp_name = f"coil_{txt_path.stem}"  # e.g. coil_10, coil_32

                    try:
                        meta, data = parse_txt_file(txt_path)
                    except Exception as exc:
                        print(f"  [ERROR] {txt_path}: {exc}", file=sys.stderr)
                        continue

                    if data.shape[0] == 0:
                        print(f"  [WARN] No data in {txt_path}", file=sys.stderr)
                        continue

                    coil_grp = sample_grp.require_group(grp_name)
                    set_attrs(coil_grp,
                              coil_current_A=I_coil,
                              source_file=txt_path.name,
                              source_path=str(txt_path.relative_to(ROOT)))

                    # Identify time, current, voltage columns from DEF metadata
                    cols = meta["columns"]
                    time_col    = next((i for i, c in enumerate(cols)
                                        if "t" == c["varname"] or "Time" == c["label"]), 0)
                    current_col = next((i for i, c in enumerate(cols)
                                        if "I" in c["varname"]), 1)
                    voltage_col = next((i for i, c in enumerate(cols)
                                        if "U" in c["varname"]), 2)

                    write_dataset(coil_grp, "time_s",
                                  data[:, time_col],
                                  unit="s", description="Elapsed time")
                    write_dataset(coil_grp, "sample_current_A",
                                  data[:, current_col],
                                  unit="A",
                                  description="Sample current (I_A1)")
                    write_dataset(coil_grp, "voltage_V",
                                  data[:, voltage_col],
                                  unit="V",
                                  description=f"{meas_type.capitalize()} voltage (U_B1)")

                    # Store header metadata as attributes
                    for meta_key in ("MIN", "MAX", "SCALE", "DEC"):
                        vals = meta.get(meta_key, [])
                        for i, col in enumerate(cols):
                            attr_name = f"header_{meta_key}_{col['varname']}"
                            if i < len(vals):
                                try:
                                    coil_grp.attrs[attr_name] = float(vals[i])
                                except ValueError:
                                    coil_grp.attrs[attr_name] = vals[i]

                    print(f"  [OK] {meas_type}/{sample_name}/{grp_name}  "
                          f"I_coil={I_coil:.1f} A  "
                          f"N={data.shape[0]}")

    print(f"\nWrote: {VOLTAGE_H5}")


# ---------------------------------------------------------------------------
# Calibration HDF5 builder (coil current -> B field)
# ---------------------------------------------------------------------------

def _try_import_openpyxl():
    try:
        import openpyxl
        return openpyxl
    except ImportError:
        return None


def _try_import_pandas():
    try:
        import pandas as pd
        return pd
    except ImportError:
        return None


def parse_calibration_excel(excel_path: Path) -> dict[str, np.ndarray]:
    """
    Read current-field.xlsx.

    Expected format: two sheets named 'increasing' and 'decreasing',
    each with a header row followed by (I(A), B(mT)) columns.

    Returns
    -------
    dict mapping sheet name -> dict with keys 'coil_current_A', 'B_field_mT'
    """
    openpyxl = _try_import_openpyxl()
    if openpyxl is None:
        raise ImportError("Install 'openpyxl': pip install openpyxl")

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    result = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        # Skip header row(s): keep only rows where first cell is numeric
        data_rows = [r for r in rows if isinstance(r[0], (int, float))]
        arr = np.array([[float(c) for c in r[:2]] for r in data_rows], dtype=float)
        result[sheet_name] = {
            "coil_current_A": arr[:, 0],
            "B_field_mT":     arr[:, 1],
        }
    wb.close()
    return result


def build_calibration_h5():
    """Convert current-field.xlsx into calibration.h5."""

    excel_path = RAW_DIR / "current-field.xlsx"

    if not excel_path.exists():
        print(f"[SKIP] {excel_path} not found.", file=sys.stderr)
        return

    HDF5_DIR.mkdir(parents=True, exist_ok=True)

    with h5py.File(CALIB_H5, "w") as hf:
        set_attrs(hf,
                  description="Coil current to magnetic field calibration",
                  experiment="Hall effect",
                  source_file=excel_path.name)

        calib_grp = hf.require_group("current_field")
        set_attrs(calib_grp,
                  description=(
                      "Magnetic field (mT) as a function of coil current (A). "
                      "Two sheets: increasing and decreasing coil current."
                  ))

        print(f"  Reading {excel_path.name} …")
        sheets = parse_calibration_excel(excel_path)

        for sheet_name, arrays in sheets.items():
            grp = calib_grp.require_group(sheet_name)
            set_attrs(grp, sweep_direction=sheet_name)
            write_dataset(grp, "coil_current_A", arrays["coil_current_A"],
                          unit="A", description="Coil (electromagnet) current")
            write_dataset(grp, "B_field_mT", arrays["B_field_mT"],
                          unit="mT", description="Magnetic field at sample position")
            print(f"    {sheet_name}: {len(arrays['coil_current_A'])} points")

    print(f"\nWrote: {CALIB_H5}")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main():
    print("=" * 60)
    print("Building voltage_measurements.h5 …")
    print("=" * 60)
    build_voltage_h5()

    print()
    print("=" * 60)
    print("Building calibration.h5 …")
    print("=" * 60)
    build_calibration_h5()


if __name__ == "__main__":
    main()
